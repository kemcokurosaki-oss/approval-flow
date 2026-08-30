# -*- coding: utf-8 -*-
import glob
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

path = glob.glob("修正内容一覧_20260827-0828.xlsx")[0]
wb = openpyxl.load_workbook(path)
ws = wb.worksheets[0]

# 既存内容をクリア
for row in ws.iter_rows():
    for cell in row:
        cell.value = None
        cell.fill = PatternFill(fill_type=None)
ws.merged_cells.ranges = list(ws.merged_cells.ranges)
for rng in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(rng))

DATA = [
    ("新機能追加", [
        ("電装（電気艤装）承認フローの新設",
         "組立フローとは別に「電装」承認フローを追加。電気艤装がある機械は組立・電装の両方が承認されて初めて完了扱いになる。承認者は組立部長のみの単一ステップ。専用チェックシート（全17項目）・電装担当ロール・メール通知も新設。"),
        ("ユニット単位承認機能の新設（2000番台）",
         "1機械に複数ユニットがある2000番台案件で、ユニットごとに個別申請・承認できる新機能。一覧モーダルで各ユニットの申請状況を表示し、行を選んでまとめて確定する操作に統一。"),
        ("分割出荷への対応",
         "1機械で工場出荷が2回に分かれるケースに対応。確定出荷日を①②の2件入力できるようにし、進捗表示も出荷ごとに分けて表示。"),
        ("3T番／4T番の承認フロー対象化",
         "従来は対象外だった点検系工番（3T／4T）も、機械組立タスクがある案件のみD番と同様に承認フロー対象に追加。工番種別フィルタ・運用ガイドも更新。"),
    ]),
    ("業務フロー簡素化", [
        ("仮出荷予定日機能の廃止",
         "出荷確定前に営業へ仮予定日を入力させていた2段階の仕組みを廃止し、確定出荷日の入力のみに簡素化。関連画面・通知・運用ガイドの文言も整理。"),
    ]),
    ("不具合修正・仕様調整", [
        ("「自分の工番」フィルタの担当判定を修正",
         "営業担当の判定基準を、工程表タスクのowner欄（実担当とズレる場合あり）から正式な工番別担当マスタに変更し、判定のズレを解消。"),
        ("チェックシート一括「○」ボタンの動作修正",
         "全項目が既に○済みのグループでも、ボタン1つで一括解除できるようトグル動作に変更。×・―が付いた項目は上書きしない。"),
    ]),
    ("表示・コード整理", [
        ("承認ステータス表示文言の統一",
         "「課長承認待ち」「部長承認待ち」等の役職別表記を「承認待ち」に統一し、表示ロジックを共通関数化。電装フロー追加に伴う表示の一貫性を確保。"),
    ]),
    ("その他（開発中の一時対応）", [
        ("デバッグ用一時ファイルの追加・削除",
         "実装内容の検証用に一時スクリプトを作成し、確認後に削除。機能への影響はなし。"),
    ]),
]

thin = Side(style="thin", color="B0B0B0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
cat_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

# タイトル
ws["A1"] = "承認フローアプリ 修正内容一覧（2026/08/27〜2026/08/28）"
ws["A1"].font = Font(size=14, bold=True)
ws.merge_cells("A1:D1")

# ヘッダー
headers = ["No", "分類", "項目", "概要"]
header_row = 3
for col, text in enumerate(headers, start=1):
    c = ws.cell(row=header_row, column=col, value=text)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

row = header_row + 1
no = 1
for category, items in DATA:
    cat_start = row
    for title, summary in items:
        ws.cell(row=row, column=1, value=no).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=2, value=category)
        ws.cell(row=row, column=3, value=title).alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row=row, column=4, value=summary).alignment = Alignment(vertical="center", wrap_text=True)
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        no += 1
    cat_end = row - 1
    if cat_end > cat_start:
        ws.merge_cells(start_row=cat_start, start_column=2, end_row=cat_end, end_column=2)
    cat_cell = ws.cell(row=cat_start, column=2)
    cat_cell.font = Font(bold=True)
    cat_cell.fill = cat_fill
    cat_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

note_row = row + 1
ws.cell(row=note_row, column=1, value="※ 「Ganttインライン編集の見た目フリーズ」バグに関する修正は今回の期間内には含まれていません。")
ws.cell(row=note_row, column=1).font = Font(italic=True, size=9, color="808080")
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)

# 列幅
widths = {"A": 6, "B": 16, "C": 30, "D": 70}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# 行の高さ(概要行は自動改行に合わせて少し高めに)
for r in range(header_row + 1, note_row):
    ws.row_dimensions[r].height = 60

ws.freeze_panes = "A4"

wb.save(path)
print("done")
