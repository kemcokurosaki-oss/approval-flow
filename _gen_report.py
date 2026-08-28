# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

rows = [
    {
        "no": 1,
        "category": "電装（電気艤装）承認フローの新規追加",
        "summary": (
            "組立フローと並行する「電装」フローを新設。電気艤装タスクがある機械は組立・電装の両方が承認されて"
            "初めて完了扱いになるよう変更。承認者は組立部長のみの単一ステップ承認（課長相当なし）。"
            "担当ロール「電装担当者（staff_denki）」を追加し、部署順一覧にも「電装」を追加。組立の丸アイコンを"
            "クリックすると組立・電装両方の状況をまとめた合成詳細モーダルが開き、個別に承認・却下できる。"
            "電気艤装完了チェックシート denki_sheet.html（機器取付／配線／艤装完了時の3グループ・計17項目＋"
            "ペンディング項目リスト）を新規作成。メール通知にも「電装」ラベル・電装完了通知本文を追加。"
        ),
        "files": "app.js, denki_sheet.html（新規）, index.html, scripts/notify-approval.js",
        "period": "08/28 12:00〜12:05、13:02〜13:18、13:56〜14:44、14:44〜14:53",
    },
    {
        "no": 2,
        "category": "2000番台「ユニット単位承認」機能の新規追加",
        "summary": (
            "1機械に複数ユニット（例：BD等）がある2000番台案件で、組立フローをユニットごとに個別申請・"
            "個別承認できるようにする新機能。ユニット一覧モーダルから各ユニットの申請状況（未申請／入力中／"
            "承認待ち／承認済み等）を一覧表示。フロー丸のアイコンは全ユニットの集約状態で色分け表示。"
            "未承認・未申請の督促バッジもユニット単位で判定するよう拡張。UIは当初「行ごとにボタン配置」だったが、"
            "最終的に「行をクリックして選択→フッターの単一ボタンで確定」する方式に変更。"
        ),
        "files": "app.js, style.css",
        "period": "08/28 14:40〜14:53、15:22〜15:29",
    },
    {
        "no": 3,
        "category": "仮出荷予定日機能の廃止（出荷日入力フローの簡素化）",
        "summary": (
            "従来は簡易検査／外観検査の完了時に営業へ「仮出荷予定日」の入力を依頼し、品証・製管が確認・確定"
            "してから出荷準備へ進む2段階の仕組みだったが、これを廃止。今後は出荷確定申請時の「確定出荷日」"
            "入力のみとし、それまでは工程表上の工場出荷タスクの予定日をそのまま表示する仕様に簡素化。"
            "関連DB列参照・通知種別・UIを削除し、運用ガイド・通知一覧・社内説明会用カンペの文言も修正。"
            "梱包出荷「未入力」時の表示文言も「梱包出荷日：未定」→「梱包出荷：あり（未入力）」に変更。"
        ),
        "files": "app.js, scripts/notify-approval.js, docs/notifications.html, guide.html, 説明会_発表者カンペ.html",
        "period": "08/27 14:45〜14:52",
    },
    {
        "no": 4,
        "category": "分割出荷（1機械で工場出荷が2回に分かれるケース）への対応",
        "summary": (
            "1つの機械に対して工程表上の「工場出荷」タスクが2件ある場合（分割出荷）に対応。承認レコードに"
            "confirmed_shipping_date_2 列を追加し、営業の確定出荷日入力フッター・変更フッターに①②2つの"
            "入力欄を出し分け。進捗カード側でも①②それぞれの出荷日を個別表示するロジックを追加。"
        ),
        "files": "app.js",
        "period": "08/28 12:57〜13:08",
    },
    {
        "no": 5,
        "category": "3T番／4T番（点検系工番）の承認フロー対象化",
        "summary": (
            "従来は承認フロー対象外だった点検系工番（3T/4T）を、D番と同様「機械組立タスクがある工番のみ」"
            "対象に含めるよう変更。進捗画面の工番種別フィルタボタンに「3T番」「4T番」を追加し、運用ガイドの"
            "説明文・工番種別一覧の記載も更新。"
        ),
        "files": "app.js, index.html, guide.html",
        "period": "08/28 09:51、10:21",
    },
    {
        "no": 6,
        "category": "「自分の工番」フィルタの担当判定を営業担当マスタ基準に修正",
        "summary": (
            "進捗一覧の「自分の担当」フィルタで、営業部については工程表タスクのowner欄（代理対応等で実担当と"
            "ズレることがある）ではなく、正式な工番別担当マスタ（app_settings.sales_person_map）を参照する"
            "よう修正。projectMatchesMine関数として共通化。"
        ),
        "files": "app.js",
        "period": "08/28 09:27",
    },
    {
        "no": 7,
        "category": "承認ステータス表示文言の統一・共通化",
        "summary": (
            "「課長承認待ち」「部長承認待ち」といった役職別の文言を「承認待ち」に統一（statusBadgeLabel関数と"
            "して一本化）。assembly／test_run／electricalの単一承認ステップ表示ロジックも共通関数化し、"
            "重複コードを整理。電装フロー追加に伴う表示の一貫性確保が目的とみられる。"
        ),
        "files": "app.js",
        "period": "08/28 14:44",
    },
    {
        "no": 8,
        "category": "機械組立/電装完了チェックシートの一括「○」ボタンの動作修正",
        "summary": (
            "sheet.html（機械組立完了チェックシート）の「全て○」ボタンについて、従来は未入力項目のみを一括で"
            "○にする仕様だったが、グループ内が既に全て○済みの場合は逆にワンクリックで一括解除できるよう修正"
            "（×や―が付いた項目は上書きしない）。トグル動作化によりチェック取り消しの手間を軽減。"
        ),
        "files": "sheet.html",
        "period": "08/27 15:24",
    },
    {
        "no": 9,
        "category": "デバッグ用一時ファイルの追加・削除（構文チェック作業）",
        "summary": (
            "機能実装の最終段階でapp.jsの構文（括弧の対応など）を検証するための一時Pythonスクリプト・JSダンプ"
            "を作成・使用後に削除。実質的な機能変更ではなく、ユニット単位承認機能などの作業中に発生した構文"
            "エラーの調査・修正作業に付随するもの。"
        ),
        "files": "（一時ファイルのみ、最終的に全て削除）",
        "period": "08/28 14:58〜15:37",
    },
]

wb = Workbook()
ws = wb.active
ws.title = "修正内容一覧"

title_font = Font(name="游ゴシック", size=14, bold=True)
header_font = Font(name="游ゴシック", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
body_font = Font(name="游ゴシック", size=10)
thin = Side(style="thin", color="B7B7B7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
wrap_center = Alignment(wrap_text=True, vertical="center", horizontal="center")

ws.merge_cells("A1:E1")
ws["A1"] = "承認フローアプリ 修正内容一覧（2026/08/27〜2026/08/28）"
ws["A1"].font = title_font
ws.row_dimensions[1].height = 24

headers = ["No", "カテゴリ", "概要", "主な変更ファイル", "日時範囲"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.border = border
    c.alignment = wrap_center

widths = [5, 32, 70, 40, 30]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

r = 4
for row in rows:
    ws.cell(row=r, column=1, value=row["no"])
    ws.cell(row=r, column=2, value=row["category"])
    ws.cell(row=r, column=3, value=row["summary"])
    ws.cell(row=r, column=4, value=row["files"])
    ws.cell(row=r, column=5, value=row["period"])
    for col in range(1, 6):
        cell = ws.cell(row=r, column=col)
        cell.font = body_font
        cell.border = border
        cell.alignment = wrap if col in (3, 4) else wrap_center
    ws.row_dimensions[r].height = 90
    r += 1

ws.freeze_panes = "A4"

note_row = r + 1
ws.merge_cells(f"A{note_row}:E{note_row}")
ws.cell(row=note_row, column=1,
        value="※ 「Ganttインライン編集の見た目フリーズ」バグに関する修正は今回の期間内には含まれていません。").font = Font(
    name="游ゴシック", size=9, italic=True, color="808080")

out_path = r"C:\Users\kurosaki\OneDrive - 日下部電機\デスクトップ\工程表作成\承認フロー\修正内容一覧_20260827-0828.xlsx"
wb.save(out_path)
print("saved:", out_path)
